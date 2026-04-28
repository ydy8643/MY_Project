import streamlit as st
import pandas as pd
from datetime import timedelta, datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import MergedCell
from io import BytesIO
import os
import holidays

st.set_page_config(layout="wide", page_title="토목1팀 근태관리 시스템")

# 데이터 저장소 초기화
if "all_data" not in st.session_state:
    st.session_state.all_data = {}

# --- [1. 보상 발생 계산 공식 함수] ---
def calculate_comp(day, night, is_holiday):
    total = day + night
    if is_holiday:
        if total <= 8:
            return total * 1.5
        else:
            return (8 * 1.5) + (total - 8) * 2.0
    else:
        # 평일: 8시간 초과분에 대해 1.5배 가산
        if total <= 8:
            return 0.0
        else:
            return (total - 8) * 1.5

# --- [2. 안전하게 엑셀 쓰기 함수] ---
def safe_write(ws, row, col, value, font=None, align=None, number_format=None):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    cell.value = value
    if font: cell.font = font
    if align: cell.alignment = align
    if number_format: cell.number_format = number_format

st.title("🏗️ 토목1팀 주간 근태 관리")

kr_holidays = holidays.KR()

# --- [3. 설정값] ---
# 실제 성함으로 변경하여 사용하세요
team_members = ["정해원", "김성호", "정종우", "김동수", "정형기", "장하나", "유두열", "최윤호"]
work_types = [
    "주간", "야간", "주간(본사)", "주간(재택)", "주간(스마트)", "주간후야간", 
    "반일후야간", "휴일근무(1)", "휴일근무(2)", "휴일야간", "오전연장(1)", 
    "오전연장(2)", "교육", "지참(0.5)", "지참(1)", "지참(1.5)", "지참(2)", 
    "지참(2.5)", "지참(3)", "지참(3.5)", "반가", "휴가", "공가", "당직후비번", 
    "주야간없음", "휴일", "기타"
]
member_row_map = {name: 5 + (i * 4) for i, name in enumerate(team_members)}

# --- [4. 날짜 설정] ---
st.sidebar.header("📅 주차 설정")
selected_date = st.sidebar.date_input("날짜 선택", value=datetime.now())
days_to_sat = (selected_date.weekday() - 5) % 7
res_start = selected_date - timedelta(days=days_to_sat)  # 이번 보고서의 '결과' 시작일 (가장 최근 토요일)
plan_start = res_start + timedelta(days=7)               # 이번 보고서의 '계획' 시작일 (다음 토요일)

res_dates = [res_start + timedelta(days=i) for i in range(7)]
plan_dates = [plan_start + timedelta(days=i) for i in range(7)]

# --- ⭐ [5. 실시간 입력 현황 대시보드 복구] ---
st.divider()
st.subheader("📊 입력 현황")
completed_names = list(st.session_state.all_data.keys())
total_count = len(team_members)
completed_count = len(completed_names)

col_m1, col_m2 = st.columns(2)
col_m1.metric("전체 인원", f"{total_count}명")
col_m2.metric("입력 완료", f"{completed_count}명", delta=f"{completed_count - total_count}명")

status_cols = st.columns(len(team_members))
for i, name in enumerate(team_members):
    with status_cols[i]:
        if name in completed_names:
            st.success(f"**{name}**\n\n✅ 완료")
        else:
            st.error(f"**{name}**\n\n⏳ 미입력")
st.divider()

# --- [6. 웹 입력 UI] ---
selected_member = st.selectbox("이름을 선택하세요", team_members)

def create_inputs(dates, key_prefix):
    day_inputs = []
    for i, d in enumerate(dates):
        is_h = d.weekday() >= 5 or (d in kr_holidays) or (d.month == 5 and d.day == 1)
        day_name = ['월','화','수','목','금','토','일'][d.weekday()]
        
        st.markdown(f"**{d.strftime('%m/%d')} ({day_name})**")
        
        # 근무형태 선택
        default_idx = work_types.index("휴일") if is_h else work_types.index("주간")
        sel_val = st.selectbox("근무형태", work_types, index=default_idx, key=f"sel_{key_prefix}_{selected_member}_{i}", label_visibility="collapsed")
        
        # 기본값 설정
        d_val, n_val = 8.0, 0.0
        if sel_val == "휴일": d_val = 0.0
        elif "지참(1)" in sel_val: d_val = 7.0
        elif "지참(1.5)" in sel_val: d_val = 6.5
        elif "반가" in sel_val: d_val = 4.0
        
        # '기타' 선택 시 수동 입력창 (날짜 하단에 배치)
        if sel_val == "기타":
            c1, c2, c3 = st.columns([1, 1, 2])
            with c1: d_val = st.number_input("주간(h)", 0.0, 24.0, 8.0, 0.5, key=f"d_{key_prefix}_{selected_member}_{i}")
            with c2: n_val = st.number_input("야간(h)", 0.0, 24.0, 0.0, 0.5, key=f"n_{key_prefix}_{selected_member}_{i}")
            with c3:
                comp_val = calculate_comp(d_val, n_val, is_h)
                st.info(f"계산된 보상: **{comp_val}**")
        else:
            comp_val = calculate_comp(d_val, n_val, is_h)
        
        day_inputs.append({"type": sel_val, "day": d_val, "night": n_val, "comp": comp_val})
        st.write("") 
        
    return day_inputs

col_res, col_plan = st.columns(2)
with col_res:
    st.subheader("⬅️ 결과 주차")
    res_data = create_inputs(res_dates, "res")
with col_plan:
    st.subheader("➡️ 계획 주차")
    plan_data = create_inputs(plan_dates, "plan")

if st.button(f"💾 {selected_member}님 데이터 저장"):
    st.session_state.all_data[selected_member] = {'res': res_data, 'plan': plan_data}
    st.toast(f"{selected_member}님 저장 완료!", icon="🎉")
    st.rerun()

# --- [7. 엑셀 추출] ---
if len(st.session_state.all_data) > 0:
    st.header("📤 엑셀 파일 생성")
    if st.button("🚀 최종 엑셀 추출"):
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            template_path = os.path.join(current_dir, 'template.xlsx')
            wb = openpyxl.load_workbook(template_path)
            
            f_blue = Font(color="0000FF", bold=True)
            f_red = Font(color="FF0000", bold=True)
            f_black = Font(color="000000", bold=True)
            align_center = Alignment(horizontal='center', vertical='center')

            def fill_ws(ws, dates, data_key):
                # 2행 M2 주차 정보 입력
                start_d, end_d = dates[0], dates[-1]
                week_num = (start_d.day - 1) // 7 + 1
                week_text = f"{start_d.year}년 {start_d.month}월 {week_num}주차({start_d.strftime('%m/%d')}~{end_d.strftime('%m/%d')})"
                safe_write(ws, 2, 13, week_text, align=Alignment(horizontal='right'))

                # 날짜 및 요일 (3-4행)
                for i, d in enumerate(dates):
                    col = 4 + i
                    font = f_red if (d.weekday() == 6 or d in kr_holidays or (d.month == 5 and d.day == 1)) else (f_blue if d.weekday() == 5 else f_black)
                    safe_write(ws, 3, col, d, font=font, align=align_center, number_format='m"/"d')
                    safe_write(ws, 4, col, ['월','화','수','목','금','토','일'][d.weekday()], font=font, align=align_center)

                # 데이터 입력
                for name, u_data in st.session_state.all_data.items():
                    r = member_row_map[name]
                    # A열 이름 병합 및 쓰기
                    safe_write(ws, r, 1, name, align=align_center)
                    try: ws.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=1)
                    except: pass
                    
                    # B-C열 양식 채우기
                    safe_write(ws, r, 2, "근무형태", align=align_center)
                    safe_write(ws, r+1, 2, "근무\n시간",
                               align=Alignment(horizontal='center', vertical='center', wrap_text=True))
                    safe_write(ws, r+1, 3, "주간", align=align_center)
                    safe_write(ws, r+2, 3, "야간", align=align_center)
                    safe_write(ws, r+3, 2, "보상발생", align=align_center)

                    for i, info in enumerate(u_data[data_key]):
                        col = 4 + i
                        # 1. 근무형태(Type)은 무조건 입력합니다. (그래야 엑셀 VLOOKUP이 이걸 보고 값을 찾아옵니다)
                        safe_write(ws, r, col, info['type'], align=align_center)
            
                        # 2. 오직 '기타'인 경우에만 파이썬이 계산한 값을 직접 입력합니다.
                        # '기타'가 아니면 아래 3줄은 실행되지 않으므로 엑셀의 VLOOKUP 수식이 그대로 살아있게 됩니다.
                        if info['type'] == "기타":
                            safe_write(ws, r+1, col, info['day'], align=align_center)
                            safe_write(ws, r+2, col, info['night'], align=align_center)
                            safe_write(ws, r+3, col, info['comp'], align=align_center)

            fill_ws(wb.worksheets[0], res_dates, 'res')
            fill_ws(wb.worksheets[1], plan_dates, 'plan')

            # 시트명 변경
            wb.worksheets[0].title = f"결과({res_dates[0].strftime('%m.%d')}~{res_dates[-1].strftime('%m.%d')})"
            wb.worksheets[1].title = f"계획({plan_dates[0].strftime('%m.%d')}~{plan_dates[-1].strftime('%m.%d')})"

            output = BytesIO()
            wb.save(output)
            st.download_button(
                label="📥 최종 엑셀 파일 받기",
                data=output.getvalue(),
                file_name=f"토목1팀 주간근무계획({plan_start.month}월 {(plan_start.day - 1) // 7 + 1}주차).xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.balloons()
        except Exception as e:
            st.error(f"오류: {e}")

# 초기화 버튼
if st.sidebar.button("🗑️ 모든 데이터 초기화"):
    st.session_state.all_data = {}
    st.rerun()